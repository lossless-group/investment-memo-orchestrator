"""
Competitive Analysis Extractor

Extracts competitive landscape data from battlecards and competitive analysis documents.
Supports multi-document synthesis for datarooms with multiple competitor battlecards.
"""

import json
import re
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime

from ..dataroom_state import CompetitiveData, CompetitorEntry, SWOTAnalysis


def extract_competitive_data(
    documents: List[Dict[str, Any]],
    use_llm: bool = True
) -> CompetitiveData:
    """
    Extract and synthesize competitive data from multiple documents.

    Args:
        documents: List of DocumentInventoryItem dicts with file_path, filename, etc.
        use_llm: Whether to use LLM for extraction (required for PDF parsing)

    Returns:
        CompetitiveData with synthesized competitive landscape
    """
    if not documents:
        return _empty_competitive_data()

    # Extract from each document
    extractions = []
    for doc in documents:
        file_path = Path(doc["file_path"])

        if file_path.suffix.lower() == ".pdf":
            extraction = extract_from_pdf(file_path, use_llm=use_llm)
        elif file_path.suffix.lower() in [".xlsx", ".xls"]:
            extraction = extract_from_excel(file_path)
        else:
            extraction = extract_generic(file_path)

        if extraction:
            extraction["source_file"] = doc["filename"]
            extractions.append(extraction)

    # Synthesize if multiple documents
    if len(extractions) == 0:
        return _empty_competitive_data()
    elif len(extractions) == 1:
        return _extraction_to_competitive_data(extractions[0])
    else:
        return synthesize_battlecards(extractions)


def extract_from_pdf(file_path: Path, use_llm: bool = True) -> Optional[Dict[str, Any]]:
    """
    Extract competitive data from PDF battlecard or analysis document.

    Args:
        file_path: Path to PDF file
        use_llm: Whether to use LLM for intelligent extraction

    Returns:
        Dict with extracted competitive data or None
    """
    try:
        from pypdf import PdfReader

        reader = PdfReader(str(file_path))

        # Extract text from all pages
        full_text = ""
        for page in reader.pages:
            text = page.extract_text()
            if text:
                full_text += f"\n{text}"

        if not full_text.strip():
            return {"extraction_notes": ["No text extracted from PDF"]}

        # Detect if this is a battlecard (competitor-specific) or general analysis
        is_battlecard = _detect_battlecard(file_path.name, full_text)

        if use_llm:
            if is_battlecard:
                return _extract_battlecard_with_llm(file_path.name, full_text)
            else:
                return _extract_analysis_with_llm(file_path.name, full_text)
        else:
            # Heuristic extraction without LLM
            return _extract_heuristic(file_path.name, full_text, is_battlecard)

    except Exception as e:
        return {"extraction_notes": [f"PDF extraction error: {str(e)}"]}


def _detect_battlecard(filename: str, content: str) -> bool:
    """Detect if document is a per-competitor battlecard."""
    filename_lower = filename.lower()

    # Filename patterns indicating battlecard
    battlecard_patterns = ["battlecard", "battle card", "vs ", "versus"]
    if any(p in filename_lower for p in battlecard_patterns):
        return True

    # Content patterns indicating battlecard
    battlecard_signals = [
        "winning angle",
        "discovery questions",
        "head-to-head",
        "our advantage",
        "their weakness",
        "feature comparison"
    ]
    content_lower = content.lower()
    signal_count = sum(1 for s in battlecard_signals if s in content_lower)

    return signal_count >= 2


def _extract_battlecard_with_llm(filename: str, content: str) -> Dict[str, Any]:
    """Use LLM to extract structured data from battlecard."""
    try:
        from langchain_anthropic import ChatAnthropic

        llm = ChatAnthropic(
            model="claude-sonnet-4-5-20250929",
            temperature=0,
            max_tokens=4000
        )

        # Extract competitor name from filename
        competitor_name = _extract_competitor_from_filename(filename)

        prompt = f"""Analyze this competitive battlecard document and extract structured data.

FILENAME: {filename}
COMPETITOR NAME (from filename): {competitor_name or "Unknown"}

DOCUMENT CONTENT:
---
{content[:12000]}
---

Extract the following as JSON:
{{
    "competitor_name": "Name of the competitor being analyzed",
    "competitor_description": "Brief description of what they do",
    "competitor_website": "Website URL if mentioned",
    "competitor_funding": "Funding amount if mentioned (number only, in USD)",
    "competitor_employees": "Employee count if mentioned (number only)",

    "strengths": ["List of competitor's strengths"],
    "weaknesses": ["List of competitor's weaknesses"],
    "threat_level": "High/Medium/Low based on content tone",

    "feature_comparison": {{
        "feature_name": {{
            "us": "Our capability (e.g., 'Full support', 'Yes', 'Native')",
            "them": "Their capability (e.g., 'Partial', 'No', 'Limited')"
        }}
    }},

    "winning_angles": ["Sales talking points for why we win"],
    "discovery_questions": ["Questions to ask prospects to expose competitor weaknesses"],

    "key_differentiators": ["Our main advantages over this competitor"],

    "extraction_notes": ["Any important observations about data quality"]
}}

Rules:
- Only include information explicitly stated in the document
- Use null for missing values, not guesses
- For feature_comparison, extract the actual comparison matrix if present
- winning_angles should be verbatim from "Our Winning Angle" or similar sections
- discovery_questions should be verbatim from "Discovery Questions" sections
"""

        response = llm.invoke(prompt)

        # Parse JSON from response
        result = _parse_json_response(response.content)

        if result:
            result["source_type"] = "battlecard"
            return result
        else:
            return {
                "competitor_name": competitor_name,
                "extraction_notes": ["Failed to parse LLM response"]
            }

    except Exception as e:
        return {
            "competitor_name": _extract_competitor_from_filename(filename),
            "extraction_notes": [f"LLM extraction error: {str(e)}"]
        }


def _extract_analysis_with_llm(filename: str, content: str) -> Dict[str, Any]:
    """Use LLM to extract data from general competitive analysis document."""
    try:
        from langchain_anthropic import ChatAnthropic

        llm = ChatAnthropic(
            model="claude-sonnet-4-5-20250929",
            temperature=0,
            max_tokens=4000
        )

        prompt = f"""Analyze this competitive analysis document and extract structured data.

FILENAME: {filename}

DOCUMENT CONTENT:
---
{content[:12000]}
---

Extract the following as JSON:
{{
    "competitors": [
        {{
            "name": "Competitor name",
            "description": "What they do",
            "strengths": ["Their strengths"],
            "weaknesses": ["Their weaknesses"],
            "threat_level": "High/Medium/Low"
        }}
    ],

    "market_positioning": "Our market position description",
    "target_segments": ["Target market segments"],

    "key_differentiators": ["Our main competitive advantages"],
    "competitive_advantages": ["Specific advantages we have"],
    "competitive_disadvantages": ["Areas where we're weaker"],

    "feature_matrix": {{
        "feature_name": {{
            "Company1": "value",
            "Company2": "value"
        }}
    }},

    "pricing_strategy": "Our pricing approach (Premium/Value/Freemium/etc)",

    "barriers_to_entry": ["Market barriers"],
    "switching_costs": "Description of customer switching costs",

    "swot": {{
        "strengths": ["Our strengths"],
        "weaknesses": ["Our weaknesses"],
        "opportunities": ["Market opportunities"],
        "threats": ["Market threats"]
    }},

    "extraction_notes": ["Any important observations"]
}}

Rules:
- Only include information explicitly stated in the document
- Use null for missing sections
- Use empty arrays [] for missing lists
"""

        response = llm.invoke(prompt)
        result = _parse_json_response(response.content)

        if result:
            result["source_type"] = "analysis"
            return result
        else:
            return {"extraction_notes": ["Failed to parse LLM response"]}

    except Exception as e:
        return {"extraction_notes": [f"LLM extraction error: {str(e)}"]}


def _extract_heuristic(filename: str, content: str, is_battlecard: bool) -> Dict[str, Any]:
    """Extract competitive data using regex patterns (no LLM)."""
    result = {
        "extraction_notes": ["Heuristic extraction (no LLM)"],
        "source_type": "battlecard" if is_battlecard else "analysis"
    }

    if is_battlecard:
        result["competitor_name"] = _extract_competitor_from_filename(filename)

    # Look for common patterns
    content_lower = content.lower()

    # Extract strengths (common patterns)
    strengths = []
    if "strength" in content_lower:
        # Try to find bullet points after "strengths"
        match = re.search(r'strengths?[:\s]+([^\n]+(?:\n[•\-\*][^\n]+)*)', content, re.I)
        if match:
            strengths = [s.strip() for s in re.split(r'[•\-\*\n]', match.group(1)) if s.strip()]
    result["strengths"] = strengths[:5]  # Limit

    # Extract weaknesses
    weaknesses = []
    if "weakness" in content_lower:
        match = re.search(r'weaknesses?[:\s]+([^\n]+(?:\n[•\-\*][^\n]+)*)', content, re.I)
        if match:
            weaknesses = [s.strip() for s in re.split(r'[•\-\*\n]', match.group(1)) if s.strip()]
    result["weaknesses"] = weaknesses[:5]

    return result


def _extract_competitor_from_filename(filename: str) -> Optional[str]:
    """Extract competitor name from battlecard filename."""
    # Common patterns:
    # "4.2 SphereBattleCard.pdf" -> "Sphere"
    # "Okta BattleCard.pdf" -> "Okta"
    # "vs Silverfort.pdf" -> "Silverfort"

    name = filename.lower()

    # Remove common suffixes
    for suffix in [".pdf", ".docx", ".xlsx", "battlecard", "battle card", "battle_card"]:
        name = name.replace(suffix, "")

    # Remove numbering prefix (e.g., "4.2 ")
    name = re.sub(r'^[\d\.\s]+', '', name)

    # Remove "vs" prefix
    name = re.sub(r'^vs\.?\s*', '', name, flags=re.I)

    # Clean and title case
    name = name.strip()
    if name:
        # Handle CamelCase (e.g., "SphereCo" -> "SphereCo")
        return name.title().replace(" ", "")

    return None


def extract_from_excel(file_path: Path) -> Optional[Dict[str, Any]]:
    """Extract competitive data from Excel comparison matrix."""
    try:
        import openpyxl

        wb = openpyxl.load_workbook(str(file_path), data_only=True)

        result = {
            "source_type": "excel_matrix",
            "feature_matrix": {},
            "competitors": [],
            "extraction_notes": []
        }

        # Look for comparison/matrix sheets
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]

            # Try to detect header row with company names
            for row_idx in range(1, min(5, sheet.max_row + 1)):
                row_values = [
                    cell.value for cell in sheet[row_idx]
                    if cell.value
                ]

                # If row has multiple non-empty cells, might be header
                if len(row_values) >= 3:
                    # First column is usually feature name
                    companies = row_values[1:]  # Skip first column

                    # Extract features from subsequent rows
                    for feat_row in range(row_idx + 1, min(50, sheet.max_row + 1)):
                        feature_name = sheet.cell(row=feat_row, column=1).value
                        if feature_name:
                            result["feature_matrix"][str(feature_name)] = {}
                            for col_idx, company in enumerate(companies, start=2):
                                value = sheet.cell(row=feat_row, column=col_idx).value
                                if value is not None:
                                    result["feature_matrix"][str(feature_name)][str(company)] = str(value)

                    result["competitors"] = [{"name": str(c)} for c in companies if c]
                    break

        wb.close()
        return result

    except Exception as e:
        return {"extraction_notes": [f"Excel extraction error: {str(e)}"]}


def extract_generic(file_path: Path) -> Optional[Dict[str, Any]]:
    """Generic extraction for unsupported formats."""
    return {
        "extraction_notes": [f"No specialized extractor for {file_path.suffix}"],
        "source_file": file_path.name
    }


def synthesize_battlecards(extractions: List[Dict[str, Any]]) -> CompetitiveData:
    """
    Synthesize multiple per-competitor battlecards into unified competitive landscape.

    This handles the common case where a dataroom contains one battlecard per competitor
    (e.g., 8 battlecards for 8 different competitors).
    """
    # Separate battlecards from general analysis docs
    battlecards = [e for e in extractions if e.get("source_type") == "battlecard"]
    analyses = [e for e in extractions if e.get("source_type") != "battlecard"]

    # Build competitor list from battlecards
    competitors: List[CompetitorEntry] = []
    all_winning_angles = []
    all_discovery_questions = []
    unified_feature_matrix: Dict[str, Dict[str, Any]] = {}

    for bc in battlecards:
        competitor = CompetitorEntry(
            name=bc.get("competitor_name", "Unknown"),
            description=bc.get("competitor_description"),
            website=bc.get("competitor_website"),
            funding_raised=_parse_funding(bc.get("competitor_funding")),
            estimated_revenue=None,
            employee_count=_parse_int(bc.get("competitor_employees")),
            founded_year=None,
            headquarters=None,
            key_customers=[],
            strengths=bc.get("strengths", []),
            weaknesses=bc.get("weaknesses", []),
            threat_level=bc.get("threat_level")
        )
        competitors.append(competitor)

        # Collect winning angles and questions
        all_winning_angles.extend(bc.get("winning_angles", []))
        all_discovery_questions.extend(bc.get("discovery_questions", []))

        # Merge feature comparisons into unified matrix
        for feature, comparison in bc.get("feature_comparison", {}).items():
            if feature not in unified_feature_matrix:
                unified_feature_matrix[feature] = {"Hydden": comparison.get("us", "")}

            competitor_name = bc.get("competitor_name", "Unknown")
            unified_feature_matrix[feature][competitor_name] = comparison.get("them", "")

    # Extract additional data from general analysis docs
    key_differentiators = []
    competitive_advantages = []
    competitive_disadvantages = []
    swot = None

    for analysis in analyses:
        key_differentiators.extend(analysis.get("key_differentiators", []))
        competitive_advantages.extend(analysis.get("competitive_advantages", []))
        competitive_disadvantages.extend(analysis.get("competitive_disadvantages", []))

        if analysis.get("swot"):
            swot = SWOTAnalysis(
                strengths=analysis["swot"].get("strengths", []),
                weaknesses=analysis["swot"].get("weaknesses", []),
                opportunities=analysis["swot"].get("opportunities", []),
                threats=analysis["swot"].get("threats", [])
            )

    # Also collect differentiators from battlecards
    for bc in battlecards:
        key_differentiators.extend(bc.get("key_differentiators", []))

    # Deduplicate lists
    key_differentiators = _dedupe_list(key_differentiators)
    all_winning_angles = _dedupe_list(all_winning_angles)
    all_discovery_questions = _dedupe_list(all_discovery_questions)

    # Build final CompetitiveData
    return CompetitiveData(
        document_source=", ".join(e.get("source_file", "") for e in extractions),
        analysis_date=datetime.now().strftime("%Y-%m-%d"),

        competitors=competitors,

        market_positioning=None,
        target_segments=[],
        geographic_focus=[],

        key_differentiators=key_differentiators,
        unique_value_proposition=None,
        competitive_advantages=competitive_advantages,
        competitive_disadvantages=competitive_disadvantages,

        feature_matrix=unified_feature_matrix if unified_feature_matrix else None,

        pricing_comparison=None,
        pricing_strategy=None,

        market_share_estimates=None,
        market_share_source=None,

        swot=swot,

        barriers_to_entry=[],
        switching_costs=None,
        network_effects=None,

        winning_angles=all_winning_angles,
        discovery_questions=all_discovery_questions,

        extraction_notes=[
            f"Synthesized from {len(battlecards)} battlecards and {len(analyses)} analysis docs",
            f"Identified {len(competitors)} competitors",
            f"Extracted {len(unified_feature_matrix)} features for comparison"
        ]
    )


def _extraction_to_competitive_data(extraction: Dict[str, Any]) -> CompetitiveData:
    """Convert a single extraction dict to CompetitiveData."""

    # Handle battlecard (single competitor)
    if extraction.get("source_type") == "battlecard":
        competitor = CompetitorEntry(
            name=extraction.get("competitor_name", "Unknown"),
            description=extraction.get("competitor_description"),
            website=extraction.get("competitor_website"),
            funding_raised=_parse_funding(extraction.get("competitor_funding")),
            estimated_revenue=None,
            employee_count=_parse_int(extraction.get("competitor_employees")),
            founded_year=None,
            headquarters=None,
            key_customers=[],
            strengths=extraction.get("strengths", []),
            weaknesses=extraction.get("weaknesses", []),
            threat_level=extraction.get("threat_level")
        )

        return CompetitiveData(
            document_source=extraction.get("source_file", ""),
            analysis_date=datetime.now().strftime("%Y-%m-%d"),
            competitors=[competitor],
            market_positioning=None,
            target_segments=[],
            geographic_focus=[],
            key_differentiators=extraction.get("key_differentiators", []),
            unique_value_proposition=None,
            competitive_advantages=[],
            competitive_disadvantages=[],
            feature_matrix=extraction.get("feature_comparison"),
            pricing_comparison=None,
            pricing_strategy=None,
            market_share_estimates=None,
            market_share_source=None,
            swot=None,
            barriers_to_entry=[],
            switching_costs=None,
            network_effects=None,
            winning_angles=extraction.get("winning_angles", []),
            discovery_questions=extraction.get("discovery_questions", []),
            extraction_notes=extraction.get("extraction_notes", [])
        )

    # Handle general analysis
    else:
        competitors = [
            CompetitorEntry(
                name=c.get("name", "Unknown"),
                description=c.get("description"),
                website=None,
                funding_raised=None,
                estimated_revenue=None,
                employee_count=None,
                founded_year=None,
                headquarters=None,
                key_customers=[],
                strengths=c.get("strengths", []),
                weaknesses=c.get("weaknesses", []),
                threat_level=c.get("threat_level")
            )
            for c in extraction.get("competitors", [])
        ]

        swot = None
        if extraction.get("swot"):
            swot = SWOTAnalysis(
                strengths=extraction["swot"].get("strengths", []),
                weaknesses=extraction["swot"].get("weaknesses", []),
                opportunities=extraction["swot"].get("opportunities", []),
                threats=extraction["swot"].get("threats", [])
            )

        return CompetitiveData(
            document_source=extraction.get("source_file", ""),
            analysis_date=datetime.now().strftime("%Y-%m-%d"),
            competitors=competitors,
            market_positioning=extraction.get("market_positioning"),
            target_segments=extraction.get("target_segments", []),
            geographic_focus=[],
            key_differentiators=extraction.get("key_differentiators", []),
            unique_value_proposition=None,
            competitive_advantages=extraction.get("competitive_advantages", []),
            competitive_disadvantages=extraction.get("competitive_disadvantages", []),
            feature_matrix=extraction.get("feature_matrix"),
            pricing_comparison=None,
            pricing_strategy=extraction.get("pricing_strategy"),
            market_share_estimates=None,
            market_share_source=None,
            swot=swot,
            barriers_to_entry=extraction.get("barriers_to_entry", []),
            switching_costs=extraction.get("switching_costs"),
            network_effects=None,
            winning_angles=[],
            discovery_questions=[],
            extraction_notes=extraction.get("extraction_notes", [])
        )


def _empty_competitive_data() -> CompetitiveData:
    """Return empty CompetitiveData structure."""
    return CompetitiveData(
        document_source="",
        analysis_date=None,
        competitors=[],
        market_positioning=None,
        target_segments=[],
        geographic_focus=[],
        key_differentiators=[],
        unique_value_proposition=None,
        competitive_advantages=[],
        competitive_disadvantages=[],
        feature_matrix=None,
        pricing_comparison=None,
        pricing_strategy=None,
        market_share_estimates=None,
        market_share_source=None,
        swot=None,
        barriers_to_entry=[],
        switching_costs=None,
        network_effects=None,
        winning_angles=[],
        discovery_questions=[],
        extraction_notes=["No competitive documents provided"]
    )


def _parse_json_response(content: str) -> Optional[Dict[str, Any]]:
    """Parse JSON from LLM response, handling markdown code blocks."""
    # Try direct parse
    try:
        return json.loads(content)
    except json.JSONDecodeError:
        pass

    # Try extracting from markdown code block
    json_match = re.search(r'```(?:json)?\s*([\s\S]*?)\s*```', content)
    if json_match:
        try:
            return json.loads(json_match.group(1))
        except json.JSONDecodeError:
            pass

    # Try finding JSON object pattern
    json_match = re.search(r'\{[\s\S]*\}', content)
    if json_match:
        try:
            return json.loads(json_match.group(0))
        except json.JSONDecodeError:
            pass

    return None


def _parse_funding(value: Any) -> Optional[float]:
    """Parse funding amount from various formats."""
    if value is None:
        return None

    if isinstance(value, (int, float)):
        return float(value)

    if isinstance(value, str):
        # Remove $ and commas
        cleaned = re.sub(r'[$,]', '', value)

        # Handle M/B suffixes
        multiplier = 1
        if 'b' in cleaned.lower():
            multiplier = 1_000_000_000
            cleaned = re.sub(r'[bB]', '', cleaned)
        elif 'm' in cleaned.lower():
            multiplier = 1_000_000
            cleaned = re.sub(r'[mM]', '', cleaned)

        try:
            return float(cleaned) * multiplier
        except ValueError:
            return None

    return None


def _parse_int(value: Any) -> Optional[int]:
    """Parse integer from various formats."""
    if value is None:
        return None

    if isinstance(value, int):
        return value

    if isinstance(value, float):
        return int(value)

    if isinstance(value, str):
        cleaned = re.sub(r'[,\s]', '', value)
        try:
            return int(float(cleaned))
        except ValueError:
            return None

    return None


def _dedupe_list(items: List[str]) -> List[str]:
    """Deduplicate list while preserving order."""
    seen = set()
    result = []
    for item in items:
        item_lower = item.lower().strip()
        if item_lower and item_lower not in seen:
            seen.add(item_lower)
            result.append(item.strip())
    return result
